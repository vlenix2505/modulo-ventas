# -*- coding: utf-8 -*-
"""
generar_datos.py — Fuente única de datos del prototipo ERP.

Datos INSPIRADOS (no copiados) en la distribuidora de food service de referencia:
  - Códigos de producto con formato AAA000 (3 letras + 3 dígitos), prefijo por familia.
  - Categorías en MAYÚSCULAS (carnes, lácteos, abarrotes, etc.).
  - Ventas con formato de comprobante FTxx-NNNNNNN.

Genera de forma coherente:
  - data/vendedores.xlsx
  - data/clientes.xlsx
  - data/productos.xlsx
  - data/ventas.xlsx        (hoja "Ventas" + hoja "Detalle")
  - js/data.js
  - tools/img_kw.tsv        (codigo<TAB>keyword, para descargar imágenes)

Ejecutar:  python tools/generar_datos.py
"""
from __future__ import annotations

import json
import random
from datetime import date, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

random.seed(42)

ROOT = Path(__file__).resolve().parent.parent
DATA_DIR = ROOT / "data"
JS_DIR = ROOT / "js"
TOOLS_DIR = ROOT / "tools"
DATA_DIR.mkdir(exist_ok=True)
JS_DIR.mkdir(exist_ok=True)

# --------------------------------------------------------------------------- #
# 1) VENDEDORES
# --------------------------------------------------------------------------- #
VENDEDORES = [
    {"nombre": "Juan Ramírez",   "email": "juan@empresa.com",   "password": "1234", "zona": "Lima"},
    {"nombre": "Roberto Laos",  "email": "roberto@empresa.com",  "password": "1234", "zona": "Lima"},
    {"nombre": "Carlos Ruiz",    "email": "carlos@empresa.com", "password": "1234", "zona": "Lima"},
    {"nombre": "Ana Flores",     "email": "ana@empresa.com",    "password": "1234", "zona": "Lima"},
    {"nombre": "María Gomez",  "email": "maria@empresa.com",  "password": "1234", "zona": "Lima"},
]

# --------------------------------------------------------------------------- #
# 2) PRODUCTOS  — formato de código AAA000, categorías en MAYÚSCULAS
#    (kw = palabra clave para la imagen real)
# --------------------------------------------------------------------------- #
PRODUCTOS = [
    {"codigo": "OFA001", "nombre": "Asado de Tira Importado USA kg",        "categoria": "CARNE DE RES IMPORTADA",      "unidad": "KG",   "precio": 52.00, "stock": 240,  "emoji": "🥩", "kw": "beef,meat"},
    {"codigo": "ADS010", "nombre": "Lomo Fino de Res Nacional kg",          "categoria": "CARNE DE RES NACIONAL",       "unidad": "KG",   "precio": 47.50, "stock": 180,  "emoji": "🥩", "kw": "steak,beef"},
    {"codigo": "ALB005", "nombre": "Pollo Entero Trozado kg",              "categoria": "AVES NACIONAL",               "unidad": "KG",   "precio": 12.40, "stock": 520,  "emoji": "🍗", "kw": "raw,chicken"},
    {"codigo": "AIB002", "nombre": "Pechuga de Pavo Importada kg",         "categoria": "AVES IMPORTADAS",             "unidad": "KG",   "precio": 28.00, "stock": 95,   "emoji": "🦃", "kw": "turkey,meat"},
    {"codigo": "LAH017", "nombre": "Bondiola de Cerdo kg",                 "categoria": "CARNICOS",                    "unidad": "KG",   "precio": 21.50, "stock": 140,  "emoji": "🥓", "kw": "pork,meat"},
    {"codigo": "LAA007", "nombre": "Queso Edam Bonlé 1kg",                 "categoria": "QUESOS",                      "unidad": "UND",  "precio": 38.00, "stock": 120,  "emoji": "🧀", "kw": "cheese,wheel"},
    {"codigo": "LAA021", "nombre": "Queso Parmesano Rallado 500g",         "categoria": "QUESOS",                      "unidad": "UND",  "precio": 31.50, "stock": 88,   "emoji": "🧀", "kw": "parmesan,cheese"},
    {"codigo": "LAB004", "nombre": "Leche Fresca Entera 1L",               "categoria": "LECHES FRESCAS",              "unidad": "UND",  "precio": 5.50,  "stock": 430,  "emoji": "🥛", "kw": "milk,bottle"},
    {"codigo": "LAD002", "nombre": "Leche Evaporada Gloria 400g",          "categoria": "EVAPORADAS",                  "unidad": "UND",  "precio": 4.20,  "stock": 980,  "emoji": "🥫", "kw": "evaporated,milk"},
    {"codigo": "LCC001", "nombre": "Leche Condensada Nestlé 1kg",          "categoria": "LECHE CONDENSADA",            "unidad": "UND",  "precio": 14.50, "stock": 160,  "emoji": "🥫", "kw": "condensed,milk"},
    {"codigo": "LAG012", "nombre": "Yogurt Bebible Fresa 1L",             "categoria": "YOGURTS",                     "unidad": "UND",  "precio": 6.80,  "stock": 360,  "emoji": "🥤", "kw": "yogurt,strawberry"},
    {"codigo": "LAG055", "nombre": "Yogurt Griego Natural 1kg",           "categoria": "YOGURTS",                     "unidad": "UND",  "precio": 11.00, "stock": 70,   "emoji": "🍶", "kw": "greek,yogurt"},
    {"codigo": "LAE004", "nombre": "Mantequilla Sin Sal Premium 1kg",      "categoria": "MANTEQUILLAS",                "unidad": "UND",  "precio": 22.00, "stock": 110,  "emoji": "🧈", "kw": "butter,block"},
    {"codigo": "LAF008", "nombre": "Manjar Blanco Repostero 1kg",          "categoria": "MANJARES",                    "unidad": "UND",  "precio": 13.00, "stock": 150,  "emoji": "🍮", "kw": "caramel,dulce"},
    {"codigo": "IGA003", "nombre": "Aceite Vegetal Premium 5L",            "categoria": "ACEITES Y GRASAS",            "unidad": "BIDON","precio": 42.00, "stock": 200,  "emoji": "🫗", "kw": "cooking,oil"},
    {"codigo": "OSB005", "nombre": "Aceite de Oliva Extra Virgen 1L",      "categoria": "ACEITES Y GRASAS",            "unidad": "UND",  "precio": 34.00, "stock": 75,   "emoji": "🫒", "kw": "olive,oil"},
    {"codigo": "AEA006", "nombre": "Caldo de Gallina Deshidratado 1kg",    "categoria": "SOPAS SUSTANCIAS Y CREMAS",   "unidad": "UND",  "precio": 26.00, "stock": 130,  "emoji": "🍲", "kw": "broth,soup"},
    {"codigo": "LAP002", "nombre": "Demiglace de Vino Tinto 1kg",          "categoria": "SOPAS SUSTANCIAS Y CREMAS",   "unidad": "UND",  "precio": 38.00, "stock": 48,   "emoji": "🍲", "kw": "sauce,gravy"},
    {"codigo": "GLO008", "nombre": "Filete de Atún en Aceite 1kg",         "categoria": "CONSERVAS",                   "unidad": "LATA", "precio": 24.00, "stock": 260,  "emoji": "🐟", "kw": "canned,tuna"},
    {"codigo": "OSE003", "nombre": "Durazno en Mitades 820g",              "categoria": "CONSERVAS",                   "unidad": "LATA", "precio": 9.50,  "stock": 310,  "emoji": "🍑", "kw": "canned,peach"},
    {"codigo": "OSF010", "nombre": "Pasta de Tomate 3kg",                  "categoria": "SALSAS Y PASTAS",             "unidad": "LATA", "precio": 18.00, "stock": 170,  "emoji": "🥫", "kw": "tomato,paste"},
    {"codigo": "OSF021", "nombre": "Salsa de Ostión 500ml",               "categoria": "SALSAS Y PASTAS",             "unidad": "UND",  "precio": 15.00, "stock": 90,   "emoji": "🍶", "kw": "soy,sauce"},
    {"codigo": "AJA007", "nombre": "Glutamato Monosódico 1kg",            "categoria": "CONDIMENTOS PROCESADOS",      "unidad": "UND",  "precio": 12.00, "stock": 140,  "emoji": "🧂", "kw": "seasoning,powder"},
    {"codigo": "CKA011", "nombre": "Pimienta Negra Molida 500g",          "categoria": "CONDIMENTOS SIN PROCESAR",    "unidad": "UND",  "precio": 28.00, "stock": 65,   "emoji": "🧂", "kw": "black,pepper"},
    {"codigo": "CCA004", "nombre": "Agua Mineral Sin Gas 625ml x15",      "categoria": "AGUA MINERAL",                "unidad": "PACK", "precio": 18.00, "stock": 400,  "emoji": "💧", "kw": "water,bottle"},
    {"codigo": "LAK002", "nombre": "Gaseosa Cola 3L",                     "categoria": "BEBIDAS",                     "unidad": "UND",  "precio": 8.50,  "stock": 280,  "emoji": "🥤", "kw": "cola,soda"},
    {"codigo": "CNA006", "nombre": "Cobertura Bitter 70% 2.5kg",          "categoria": "COBERTURAS CHOCOLATES Y OTROS","unidad": "UND", "precio": 64.00, "stock": 55,   "emoji": "🍫", "kw": "dark,chocolate"},
    {"codigo": "CNV003", "nombre": "Gelatina Sin Sabor 500g",             "categoria": "CONFITERIA",                  "unidad": "UND",  "precio": 22.00, "stock": 95,   "emoji": "🍮", "kw": "gelatin,dessert"},
    {"codigo": "FDA002", "nombre": "Fideo Spaghetti Don Vittorio 5kg",    "categoria": "FIDEOS",                      "unidad": "BOLSA","precio": 26.00, "stock": 230,  "emoji": "🍝", "kw": "spaghetti,pasta"},
    {"codigo": "ARA001", "nombre": "Arroz Superior Añejo 50kg",           "categoria": "ARROZ",                       "unidad": "SACO", "precio": 165.00,"stock": 120,  "emoji": "🍚", "kw": "rice,sack"},
    {"codigo": "AZA001", "nombre": "Azúcar Rubia Doméstica 50kg",         "categoria": "AZUCAR",                      "unidad": "SACO", "precio": 145.00,"stock": 110,  "emoji": "🧂", "kw": "sugar,sack"},
    {"codigo": "HVA001", "nombre": "Huevos Pardos Frescos x30",           "categoria": "HUEVOS",                      "unidad": "PACK", "precio": 16.00, "stock": 340,  "emoji": "🥚", "kw": "eggs,carton"},
    {"codigo": "MRA003", "nombre": "Langostino Entero Congelado kg",       "categoria": "MARISCOS Y CONGELADOS",       "unidad": "KG",   "precio": 38.00, "stock": 80,   "emoji": "🦐", "kw": "shrimp,seafood"},
    {"codigo": "PCA001", "nombre": "Papa Bastón Congelada 2.5kg",         "categoria": "PAPAS CONGELADAS",            "unidad": "BOLSA","precio": 19.00, "stock": 260,  "emoji": "🍟", "kw": "french,fries"},
    {"codigo": "VGA004", "nombre": "Mix de Vegetales Congelado 1kg",      "categoria": "VERDURAS CONGELADAS",         "unidad": "BOLSA","precio": 9.00,  "stock": 190,  "emoji": "🥦", "kw": "frozen,vegetables"},
    {"codigo": "LBA005", "nombre": "Pulpa de Maracuyá Congelada 1kg",     "categoria": "FRUTAS CONGELADAS",           "unidad": "BOLSA","precio": 15.00, "stock": 130,  "emoji": "🥭", "kw": "passion,fruit"},
]
for p in PRODUCTOS:
    p["imagen"] = f"data/img/{p['codigo']}.jpg"

# --------------------------------------------------------------------------- #
# 3) CLIENTES  — por vendedor: 2 HOTELES Y ENTRETENIMIENTO, 2 RESTAURANTES, 1 otro
# --------------------------------------------------------------------------- #
HOTELES = [
    ("Hotel Costa del Sol",       "5 ESTRELLAS", "CADENA"),
    ("Hotel Sonesta Posadas",     "4 ESTRELLAS", "CADENA"),
    ("Hotel Boutique Miraflores", "BOUTIQUE",    "INDEPENDIENTE"),
    ("Resort Las Dunas",          "RESORT",      "CADENA"),
    ("Casino Atlantic City",      "CASINO",      "ENTRETENIMIENTO"),
    ("Hotel La Hacienda",         "4 ESTRELLAS", "INDEPENDIENTE"),
    ("Casino Fiesta",             "CASINO",      "ENTRETENIMIENTO"),
    ("Hotel Libertador",          "5 ESTRELLAS", "CADENA"),
    ("Club Náutico Ancón",        "CLUB",        "ENTRETENIMIENTO"),
    ("Hotel Plaza Real",          "BOUTIQUE",    "INDEPENDIENTE"),
]
RESTAURANTES = [
    ("Restaurante La Mar",         "MARINA",      "GOURMET"),
    ("Pollería Norkys",            "POLLERIA",    "SALON"),
    ("Cevichería El Muelle",       "CEVICHERIA",  "MARINA"),
    ("Chifa Titi",                 "CHIFA",       "SALON"),
    ("Parrillas El Tizón",         "PARRILLA",    "GOURMET"),
    ("Restaurante Pardos",         "POLLERIA",    "CADENA"),
    ("Marisquería La Red",         "CEVICHERIA",  "MARINA"),
    ("Restaurante El Aguajal",     "CRIOLLO",     "FAMILIAR"),
    ("Parrilla Don Carbón",        "PARRILLA",    "SALON"),
    ("Cevichería La Picantería",   "CEVICHERIA",  "FAMILIAR"),
]
OTROS = [
    ("Panadería San Antonio",  "PANADERIAS Y PASTELERIAS", "ARTESANAL",   "SALON"),
    ("Cafetería Altomayo",     "CAFETERIAS",               "ESPECIALIDAD","CADENA"),
    ("Catering Eventos VIP",   "CATERING Y EVENTOS",       "CORPORATIVO", "PREMIUM"),
    ("Burger House Express",   "COMIDA RAPIDA",            "HAMBURGUESAS","FAST FOOD"),
    ("Minimarket La Canasta",  "MINIMARKETS",              "INDEPENDIENTE","BARRIO"),
]

CLIENTES = []
_n = 0
for vi, v in enumerate(VENDEDORES):
    seleccion = []
    # 2 hoteles y entretenimiento
    for nombre, sr1, sr2 in HOTELES[vi * 2: vi * 2 + 2]:
        seleccion.append((nombre, "HOTELES Y ENTRETENIMIENTO", sr1, sr2))
    # 2 restaurantes
    for nombre, sr1, sr2 in RESTAURANTES[vi * 2: vi * 2 + 2]:
        seleccion.append((nombre, "RESTAURANTES", sr1, sr2))
    # 1 de rubro variado
    nombre, rubro, sr1, sr2 = OTROS[vi]
    seleccion.append((nombre, rubro, sr1, sr2))

    for nombre, rubro, sr1, sr2 in seleccion:
        _n += 1
        CLIENTES.append({
            "id": f"CL{_n:03d}",
            "nombre": nombre,
            "rubro": rubro,
            "subrubro1": sr1,
            "subrubro2": sr2,
            "sede": v["zona"],
            "vendedorEmail": v["email"],
        })

# --------------------------------------------------------------------------- #
# 4) VENTAS HISTÓRICAS  — id estilo comprobante FTxx-NNNNNNN
# --------------------------------------------------------------------------- #
# S1: clientes con historial (≥4 compras)   S2: cold-start (≤3 compras)
_S1 = {'CL001','CL003','CL006','CL008','CL011','CL014','CL017','CL018','CL021','CL023'}
_S2 = {'CL002','CL004','CL007','CL009','CL012','CL013','CL016','CL019','CL022','CL024'}

def generar_ventas() -> list[dict]:
    ventas = []
    nro = 0
    hoy = date(2026, 6, 21)
    for cli in CLIENTES:
        if cli['id'] in _S1:
            n_ventas = random.randint(11, 15)
        elif cli['id'] in _S2:
            n_ventas = random.randint(2, 3)
        else:
            n_ventas = random.randint(2, 4)
        for _ in range(n_ventas):
            nro += 1
            dias_atras = random.randint(5, 250)
            fecha = (hoy - timedelta(days=dias_atras)).isoformat()
            items = []
            for prod in random.sample(PRODUCTOS, random.randint(3, 6)):
                cant = random.randint(1, 24)
                subtotal = round(cant * prod["precio"], 2)
                items.append({
                    "codigo": prod["codigo"], "nombre": prod["nombre"],
                    "cantidad": cant, "precio": prod["precio"], "subtotal": subtotal,
                })
            total = round(sum(i["subtotal"] for i in items), 2)
            ventas.append({
                "ventaId": f"FT0{random.choice([1, 2, 4])}-{9000000 + nro}",
                "clienteId": cli["id"], "fecha": fecha,
                "vendedorEmail": cli["vendedorEmail"], "items": items, "total": total,
            })
    ventas.sort(key=lambda v: v["fecha"])
    return ventas

VENTAS = generar_ventas()

# --------------------------------------------------------------------------- #
# Estilo Excel
# --------------------------------------------------------------------------- #
HEADER_FILL = PatternFill("solid", fgColor="0D9488")
HEADER_FONT = Font(bold=True, color="FFFFFF")

def estilar_cabecera(ws, ncols: int) -> None:
    for col in range(1, ncols + 1):
        c = ws.cell(row=1, column=col)
        c.fill = HEADER_FILL; c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 20

def autoancho(ws) -> None:
    for col in ws.columns:
        ancho = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(ancho + 4, 50)

def excel_vendedores() -> None:
    wb = Workbook(); ws = wb.active; ws.title = "Vendedores"
    ws.append(["nombre", "email", "password", "zona"])
    for v in VENDEDORES:
        ws.append([v["nombre"], v["email"], v["password"], v["zona"]])
    estilar_cabecera(ws, 4); autoancho(ws); wb.save(DATA_DIR / "vendedores.xlsx")

def excel_clientes() -> None:
    wb = Workbook(); ws = wb.active; ws.title = "Clientes"
    ws.append(["id", "nombre", "rubro", "subrubro1", "subrubro2", "sede", "vendedorEmail"])
    for c in CLIENTES:
        ws.append([c["id"], c["nombre"], c["rubro"], c["subrubro1"], c["subrubro2"], c["sede"], c["vendedorEmail"]])
    estilar_cabecera(ws, 7); autoancho(ws); wb.save(DATA_DIR / "clientes.xlsx")

def excel_productos() -> None:
    wb = Workbook(); ws = wb.active; ws.title = "Productos"
    ws.append(["codigo", "nombre", "categoria", "unidad", "precio", "stock", "emoji", "imagen"])
    for p in PRODUCTOS:
        ws.append([p["codigo"], p["nombre"], p["categoria"], p["unidad"], p["precio"], p["stock"], p["emoji"], p["imagen"]])
    estilar_cabecera(ws, 8); autoancho(ws); wb.save(DATA_DIR / "productos.xlsx")

def excel_ventas() -> None:
    wb = Workbook(); ws = wb.active; ws.title = "Ventas"
    ws.append(["ventaId", "clienteId", "fecha", "vendedorEmail", "total"])
    for v in VENTAS:
        ws.append([v["ventaId"], v["clienteId"], v["fecha"], v["vendedorEmail"], v["total"]])
    estilar_cabecera(ws, 5); autoancho(ws)

    wd = wb.create_sheet("Detalle")
    wd.append(["ventaId", "codigo", "nombre", "cantidad", "precio", "subtotal"])
    for v in VENTAS:
        for it in v["items"]:
            wd.append([v["ventaId"], it["codigo"], it["nombre"], it["cantidad"], it["precio"], it["subtotal"]])
    estilar_cabecera(wd, 6); autoancho(wd); wb.save(DATA_DIR / "ventas.xlsx")

def escribir_data_js() -> None:
    payload = {"vendedores": VENDEDORES,
               "clientes": CLIENTES,
               "productos": [{k: v for k, v in p.items() if k != "kw"} for p in PRODUCTOS],
               "ventas": VENTAS}
    js = ("// AUTOGENERADO por tools/generar_datos.py — no editar a mano.\n"
          "// Datos semilla del prototipo. La app los copia a localStorage en el primer arranque.\n"
          "window.SEED_DATA = " + json.dumps(payload, ensure_ascii=False, indent=2) + ";\n")
    (JS_DIR / "data.js").write_text(js, encoding="utf-8")

def escribir_kw() -> None:
    lineas = [f"{p['codigo']}\t{p['kw']}" for p in PRODUCTOS]
    (TOOLS_DIR / "img_kw.tsv").write_text("\n".join(lineas) + "\n", encoding="utf-8")

def main() -> None:
    excel_vendedores(); excel_clientes(); excel_productos(); excel_ventas()
    escribir_data_js(); escribir_kw()
    rub = {}
    for c in CLIENTES:
        rub[c["rubro"]] = rub.get(c["rubro"], 0) + 1
    print("OK -> data/*.xlsx, js/data.js, tools/img_kw.tsv")
    print(f"   vendedores={len(VENDEDORES)}  clientes={len(CLIENTES)}  productos={len(PRODUCTOS)}  ventas={len(VENTAS)}")
    print("   clientes por rubro:", rub)

if __name__ == "__main__":
    main()
